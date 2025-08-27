#pip install openpyxl
#pip install pandas
import openpyxl
import pandas as pd
import os
import csv
import ollama

directory_path = '.\\dados'

arquivos = [
    "ADMISSÇO ABRIL.xlsx.csv",
    "AFASTAMENTOS.xlsx.csv",
    "APRENDIZ.xlsx.csv",
    "ATIVOS.xlsx.csv",
    "Base dias uteis.xlsx.csv",
    "Base sindicato x valor.xlsx.csv",
    "DESLIGADOS.xlsx.csv",
    "ESTµGIO.xlsx.csv",
    "EXTERIOR.xlsx.csv",
    "FRIAS.xlsx.csv"
]

instrucoes = [
    "How many lines exists in following CSV spreadsheet? ",
    "Which is the sum of all 'VALOR NOTA FISCAL' column values in following CSV spreadsheet? ",
    "Which is the major and maximum value of all 'VALOR NOTA FISCAL' column in following CSV spreadsheet? ",
    "Which is the major and maximum value of all 'VALOR NOTA FISCAL' column in following CSV spreadsheet? "+
    "After that response which is the value of the 'CPF/CNPJ Emitente' of the same line that have the previous founded value (major and maximum value of all 'VALOR NOTA FISCAL' column). ",
    "For just the lines with value 378257000181 on 'CPF/CNPJ Emitente' column, sum values of 'VALOR NOTA FISCAL' column?",
    "What is the sum of all values of 'VALOR NOTA FISCAL' column?"
    "Calculate and respond these two following proposals:\n"+
    "1. Select, reserve and tell how many lines exists in the CSV with value 378257000181 on 'CPF/CNPJ Emitente' column?\n"+
    "2. Using just only the lines selected by the first proposal, filter and calcule the sum of values of 'VALOR NOTA FISCAL' column",
    "What is the value of 'CPF/CNPJ Emitente' at the second line at the original CSV content?"
]

def delete_files_by_extension(directory_path, extension):
    """
    Deletes all files with a specific extension in a given directory.

    Args:
        directory_path (str): The path to the directory.
        extension (str): The file extension to target (e.g., ".txt", ".log").
    """
    try:
        # List all items in the specified directory
        for item_name in os.listdir(directory_path):
            # Construct the full path to the item
            full_path = os.path.join(directory_path, item_name)

            # Check if the item is a file and ends with the desired extension
            if os.path.isfile(full_path) and item_name.endswith(extension):
                os.remove(full_path)
                #print(f"Deleted: {full_path}")
    except OSError as e:
        print(f"Error: {e}")

delete_files_by_extension('.\\', '.csv')

#converte os arquivos XLSX para CSV usando o Pandas:
for entry in os.scandir(directory_path):
    if entry.is_file():
        #print(entry.path)
        try:
            input_xlsx_file = entry.path
            output_csv_file = entry.path.replace(".xlsx", "").replace("\\dados", "") + ".csv"
            # Le o arquivo Excel para o pandas DataFrame
            df = pd.read_excel(input_xlsx_file)

            # Converte para CSV
            df.to_csv(output_csv_file, index=False)

            #print(f"Convertendo arquivo do excel: '{input_xlsx_file}' para '{output_csv_file}'")
            
#            for question in questions:
#                print("\n\n"+question)
#                chat(preface, cabecalho, prepare, question)

# Itera os arquivos CSV            
            with open(output_csv_file, 'r') as file:
                csv_reader = csv.reader(file)

                # Optionally, skip the header row if present
                header = next(csv_reader)
                #print(f"CREATE TABLE {file} ( {header} );")
                print(f"CREATE TABLE {file.name.replace(".\\","").replace(" ","_").replace(".xlsx.csv","")} (", end="")
                for i, item in enumerate(header):
                    print(item.replace(" ","_").replace(":","_"), end="")
                    print(" VARCHAR(50)", end="")
                    if i < len(header) - 1:
                        print(",", end="")
                print(");")
                        
                # Iterate and print each data row
                for row in csv_reader:
                    print(row)
            
# Itera os arquivos XLSX
#            # Load the workbook
#            try:
#                workbook = openpyxl.load_workbook(input_xlsx_file)
#            except FileNotFoundError:
#                print("Error: '{output_csv_file}' not found. Please ensure the file exists.")
#                exit()
#
#            # Select the active worksheet (or a specific sheet by name: workbook["Sheet1"])
#            sheet = workbook.active
#
#            # Iterate through rows and cells
#            print("Iterating through rows and cells:")
#            for row in sheet.iter_rows():  # You can specify min_row, max_row, min_col, max_col
#                for cell in row:
#                    print(cell.value, end="\t")  # Print cell value, separated by tab
#                print()  # Newline after each row
#
#            # Alternatively, iterate by row index and column index
#            print("\nIterating by row and column index:")
#            for r_idx in range(1, sheet.max_row + 1):
#                for c_idx in range(1, sheet.max_column + 1):
#                    cell_obj = sheet.cell(row=r_idx, column=c_idx)
#                    print(cell_obj.value, end="\t")
#                #print()

        except FileNotFoundError:
            print(f"Erro: Arquivo '{input_xlsx_file}' não convertido.")
        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            
#delete_files_by_extension('.\\', '.csv')